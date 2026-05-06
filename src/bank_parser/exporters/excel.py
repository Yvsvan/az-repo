"""Exportador a Excel (.xlsx).

Una hoja por RFC con las columnas estandarizadas, totales al pie,
hoja Resumen con metadatos de cada Statement y hoja Advertencias
(sólo si hay warnings).
"""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Sequence
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bank_parser.core.schema import Movement, Statement

# Columnas estandarizadas en el orden canónico
COLUMNAS: tuple[str, ...] = (
    "fecha",
    "descripcion",
    "abono",
    "cargo",
    "saldo",
    "banco",
    "cuenta",
    "archivo_origen",
)

# Columnas del resumen
_RESUMEN_COLS: tuple[str, ...] = (
    "banco",
    "rfc",
    "titular",
    "cuenta",
    "periodo_inicio",
    "periodo_fin",
    "saldo_inicial",
    "saldo_final",
    "total_abonos",
    "total_cargos",
    "num_movimientos",
    "archivo_origen",
)

# Estilos
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_TOTAL_FONT = Font(bold=True)
_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_AMOUNT_NUMBER_FORMAT = "#,##0.00"
_DATE_NUMBER_FORMAT = "YYYY-MM-DD"

_SIN_RFC_PREFIX = "SIN_RFC"


def export_to_xlsx(
    statements: Sequence[Statement],
    output_path: Path | str,
) -> Path:
    """Exporta uno o varios :class:`Statement` a un archivo ``.xlsx``.

    Args:
        statements: Lista de estados de cuenta ya parseados.
        output_path: Ruta destino del archivo Excel. Los directorios
            intermedios se crean automáticamente.

    Returns:
        El :class:`Path` del archivo generado.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Eliminar la hoja vacía que crea openpyxl por defecto
    wb.remove(wb.active)

    # Agrupar statements por RFC
    groups: dict[str, list[Statement]] = defaultdict(list)
    sin_rfc_counter = 0
    rfc_sheet_names: dict[str, str] = {}  # rfc_key -> sheet_name

    for stmt in statements:
        rfc = stmt.summary.rfc
        if rfc:
            key = rfc
        else:
            sin_rfc_counter += 1
            key = f"{_SIN_RFC_PREFIX}_{sin_rfc_counter}"
        groups[key].append(stmt)
        rfc_sheet_names[key] = key  # nombre de hoja = clave RFC

    # Crear hoja por RFC con los movimientos
    for rfc_key, stmts in groups.items():
        all_movements: list[Movement] = []
        for stmt in stmts:
            all_movements.extend(stmt.movements)

        ws = wb.create_sheet(title=rfc_key)
        _write_movements_sheet(ws, all_movements)

    # Hoja Resumen
    ws_resumen = wb.create_sheet(title="Resumen")
    _write_resumen_sheet(ws_resumen, list(statements))

    # Hoja Advertencias (sólo si algún statement tiene warnings)
    all_warnings = [(stmt, w) for stmt in statements for w in stmt.warnings]
    if all_warnings:
        ws_warn = wb.create_sheet(title="Advertencias")
        _write_warnings_sheet(ws_warn, all_warnings)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# helpers privados
# ---------------------------------------------------------------------------


def _apply_header_style(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit_columns(ws) -> None:
    """Ajusta el ancho de columnas al contenido (heurística)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _write_movements_sheet(ws, movements: list[Movement]) -> None:
    """Escribe encabezado + filas de movimientos + fila TOTAL."""
    # Encabezado
    for col, name in enumerate(COLUMNAS, 1):
        ws.cell(1, col, name)
    _apply_header_style(ws, len(COLUMNAS))

    amount_cols = {
        list(COLUMNAS).index("abono") + 1,
        list(COLUMNAS).index("cargo") + 1,
        list(COLUMNAS).index("saldo") + 1,
    }
    fecha_col = list(COLUMNAS).index("fecha") + 1

    # Filas de movimientos
    for row_idx, mov in enumerate(movements, 2):
        ws.cell(row_idx, 1, mov.fecha)
        ws.cell(row_idx, 2, mov.descripcion)
        ws.cell(row_idx, 3, float(mov.abono))
        ws.cell(row_idx, 4, float(mov.cargo))
        ws.cell(row_idx, 5, float(mov.saldo))
        ws.cell(row_idx, 6, mov.banco.value)
        ws.cell(row_idx, 7, mov.cuenta)
        ws.cell(row_idx, 8, mov.archivo_origen)

        ws.cell(row_idx, fecha_col).number_format = _DATE_NUMBER_FORMAT
        for c in amount_cols:
            ws.cell(row_idx, c).number_format = _AMOUNT_NUMBER_FORMAT

    # Fila TOTAL
    total_row = len(movements) + 2
    total_abono = sum(float(m.abono) for m in movements)
    total_cargo = sum(float(m.cargo) for m in movements)

    desc_col = list(COLUMNAS).index("descripcion") + 1
    abono_col = list(COLUMNAS).index("abono") + 1
    cargo_col = list(COLUMNAS).index("cargo") + 1

    ws.cell(total_row, desc_col, "TOTAL")
    ws.cell(total_row, abono_col, total_abono)
    ws.cell(total_row, cargo_col, total_cargo)

    for col in range(1, len(COLUMNAS) + 1):
        cell = ws.cell(total_row, col)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
    for c in (abono_col, cargo_col):
        ws.cell(total_row, c).number_format = _AMOUNT_NUMBER_FORMAT

    _autofit_columns(ws)


def _write_resumen_sheet(ws, statements: list[Statement]) -> None:
    """Escribe la hoja Resumen con una fila por Statement."""
    for col, name in enumerate(_RESUMEN_COLS, 1):
        ws.cell(1, col, name)
    _apply_header_style(ws, len(_RESUMEN_COLS))

    amount_cols_names = {"saldo_inicial", "saldo_final", "total_abonos", "total_cargos"}
    date_cols_names = {"periodo_inicio", "periodo_fin"}
    col_idx = {name: i + 1 for i, name in enumerate(_RESUMEN_COLS)}

    for row_idx, stmt in enumerate(statements, 2):
        s = stmt.summary
        values = {
            "banco": s.banco.value,
            "rfc": s.rfc or "",
            "titular": s.titular,
            "cuenta": s.cuenta,
            "periodo_inicio": s.periodo_inicio,
            "periodo_fin": s.periodo_fin,
            "saldo_inicial": float(s.saldo_inicial),
            "saldo_final": float(s.saldo_final),
            "total_abonos": float(s.total_abonos),
            "total_cargos": float(s.total_cargos),
            "num_movimientos": len(stmt.movements),
            "archivo_origen": s.archivo_origen,
        }
        for name, val in values.items():
            cell = ws.cell(row_idx, col_idx[name], val)
            if name in amount_cols_names:
                cell.number_format = _AMOUNT_NUMBER_FORMAT
            elif name in date_cols_names and val is not None:
                cell.number_format = _DATE_NUMBER_FORMAT

    _autofit_columns(ws)


def _write_warnings_sheet(ws, warnings: list[tuple[Statement, str]]) -> None:
    """Escribe la hoja Advertencias."""
    headers = ["banco", "rfc", "archivo_origen", "advertencia"]
    for col, name in enumerate(headers, 1):
        ws.cell(1, col, name)
    _apply_header_style(ws, len(headers))

    for row_idx, (stmt, warning) in enumerate(warnings, 2):
        ws.cell(row_idx, 1, stmt.summary.banco.value)
        ws.cell(row_idx, 2, stmt.summary.rfc or "")
        ws.cell(row_idx, 3, stmt.summary.archivo_origen)
        ws.cell(row_idx, 4, warning)

    _autofit_columns(ws)
