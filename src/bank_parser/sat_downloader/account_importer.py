"""Importador XLSX de cuentas SAT y generador de plantilla."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from bank_parser.sat_downloader.auth import load_credential, save_password
from bank_parser.sat_downloader.schema import SatCredential

_HEADERS = [
    "RFC",
    "Alias (opcional)",
    "Ruta Certificado (.cer)",
    "Ruta Llave Privada (.key)",
    "Contraseña",
]
_COL_WIDTHS = [22, 26, 48, 48, 22]

_EXAMPLE_ROW = [
    "XAXX010101000",
    "Mi Empresa",
    r"C:\certs\fiel.cer",
    r"C:\certs\fiel.key",
    "mi_contraseña",
]


def generate_template(dest: Path) -> None:
    """Genera una plantilla XLSX con encabezados y fila de ejemplo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuentas SAT"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=11)
    example_font = Font(name="Segoe UI", color="9E9E9E", italic=True, size=10)

    for col, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS, strict=False), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    for col, value in enumerate(_EXAMPLE_ROW, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = example_font

    # Nota de ayuda
    ws.cell(row=4, column=1, value="Notas:").font = Font(bold=True, name="Segoe UI", size=10)
    ws.cell(
        row=5,
        column=1,
        value="• La columna RFC se auto-detecta del certificado; puedes dejarla vacía.",
    ).font = Font(name="Segoe UI", size=10, color="555555")
    ws.cell(
        row=6,
        column=1,
        value="• La contraseña se almacena en Windows Credential Manager, no en este archivo.",
    ).font = Font(name="Segoe UI", size=10, color="555555")

    wb.save(dest)


def import_from_xlsx(path: Path) -> tuple[list[SatCredential], list[str]]:
    """Lee un XLSX con formato de plantilla y carga las credenciales.

    Retorna (credentials_cargadas, mensajes_de_error).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    credentials: list[SatCredential] = []
    errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        # Ignorar filas completamente vacías o de notas
        if not row or not any(row[:5]):
            continue
        cells = list(row) + [None] * 5
        _rfc, alias_val, cer_val, key_val, pwd_val = cells[:5]

        alias = str(alias_val).strip() if alias_val else ""
        cer_str = str(cer_val).strip() if cer_val else ""
        key_str = str(key_val).strip() if key_val else ""
        password = str(pwd_val).strip() if pwd_val else ""

        if not cer_str or not key_str or not password:
            errors.append(f"Fila {i}: faltan campos obligatorios (ruta .cer, .key o contraseña)")
            continue

        cer_path = Path(cer_str)
        key_path = Path(key_str)

        if not cer_path.exists():
            errors.append(f"Fila {i}: certificado no encontrado → {cer_str}")
            continue
        if not key_path.exists():
            errors.append(f"Fila {i}: llave privada no encontrada → {key_str}")
            continue

        try:
            cred = load_credential(cer_path, key_path, password, alias)
            save_password(cred.rfc, password)
            credentials.append(cred)
        except Exception as exc:
            errors.append(f"Fila {i}: error al cargar credencial — {exc}")

    return credentials, errors
