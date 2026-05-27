"""Export parsed CFDI rows to a multi-sheet Excel workbook."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bank_parser.cfdi_converter.schema import CfdiRow, NominaRow, PagoDocRow

AnyRow = CfdiRow | NominaRow | PagoDocRow

# Styles (matching the bank parser Excel exporter palette)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_TOTAL_FONT = Font(bold=True)
_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_AMOUNT_FMT = "#,##0.00"
_DATE_FMT = "YYYY-MM-DD"
_RATE_FMT = "#,##0.0000"

# Column spec: (header_label, dataclass_field, format_type)
# format_type: "amount" | "date" | "rate" | "str"
_ColSpec = tuple[str, str, str]

_FACTURA_COLS: list[_ColSpec] = [
    ("UUID", "uuid", "str"),
    ("Fecha", "fecha", "date"),
    ("Fecha Timbrado", "fecha_timbrado", "date"),
    ("Tipo", "tipo", "str"),
    ("Serie", "serie", "str"),
    ("Folio", "folio", "str"),
    ("RFC Emisor", "rfc_emisor", "str"),
    ("Nombre Emisor", "nombre_emisor", "str"),
    ("Régimen Fiscal Emisor", "regimen_fiscal_emisor", "str"),
    ("RFC Receptor", "rfc_receptor", "str"),
    ("Nombre Receptor", "nombre_receptor", "str"),
    ("Uso CFDI", "uso_cfdi", "str"),
    ("Subtotal", "subtotal", "amount"),
    ("Descuento", "descuento", "amount"),
    ("IVA Trasladado", "iva_trasladado", "amount"),
    ("IVA Retenido", "iva_retenido", "amount"),
    ("ISR Retenido", "isr_retenido", "amount"),
    ("IEPS", "ieps", "amount"),
    ("Total", "total", "amount"),
    ("Moneda", "moneda", "str"),
    ("Tipo de Cambio", "tipo_cambio", "rate"),
    ("Total MXN", "total_mxn", "amount"),
    ("Forma de Pago", "forma_pago", "str"),
    ("Método de Pago", "metodo_pago", "str"),
    ("Descripción", "descripcion", "str"),
    ("Lugar Expedición", "lugar_expedicion", "str"),
    ("Archivo XML", "archivo_xml", "str"),
]

_NOMINA_COLS: list[_ColSpec] = [
    ("UUID", "uuid", "str"),
    ("Fecha", "fecha", "date"),
    ("Fecha Timbrado", "fecha_timbrado", "date"),
    ("RFC Emisor", "rfc_emisor", "str"),
    ("Nombre Emisor", "nombre_emisor", "str"),
    ("RFC Empleado", "rfc_empleado", "str"),
    ("Nombre Empleado", "nombre_empleado", "str"),
    ("CURP", "curp", "str"),
    ("Num. Empleado", "num_empleado", "str"),
    ("Total", "total", "amount"),
    ("Tipo Nómina", "tipo_nomina", "str"),
    ("Fecha Pago", "fecha_pago", "date"),
    ("Fecha Ini. Pago", "fecha_ini_pago", "date"),
    ("Fecha Fin Pago", "fecha_fin_pago", "date"),
    ("Días Pagados", "num_dias_pagados", "amount"),
    ("Total Percepciones", "total_percepciones", "amount"),
    ("Total Deducciones", "total_deducciones", "amount"),
    ("Otros Pagos", "total_otros_pagos", "amount"),
    ("Departamento", "departamento", "str"),
    ("Puesto", "puesto", "str"),
    ("Riesgo Laboral", "riesgo_lab", "str"),
    ("Tipo Contrato", "tipo_contrato", "str"),
    ("Archivo XML", "archivo_xml", "str"),
]

_PAGO_COLS: list[_ColSpec] = [
    ("UUID Pago", "uuid_pago", "str"),
    ("Fecha Pago", "fecha_pago", "date"),
    ("RFC Emisor", "rfc_emisor", "str"),
    ("Nombre Emisor", "nombre_emisor", "str"),
    ("RFC Receptor", "rfc_receptor", "str"),
    ("Nombre Receptor", "nombre_receptor", "str"),
    ("Forma de Pago", "forma_pago", "str"),
    ("Moneda Pago", "moneda_pago", "str"),
    ("Tipo Cambio Pago", "tipo_cambio_pago", "rate"),
    ("UUID Relacionado", "uuid_relacionado", "str"),
    ("Serie Doc.", "serie_doc", "str"),
    ("Folio Doc.", "folio_doc", "str"),
    ("Método Pago Doc.", "metodo_pago_doc", "str"),
    ("Num. Parcialidad", "num_parcialidad", "str"),
    ("Saldo Anterior", "imp_saldo_ant", "amount"),
    ("Importe Pagado", "imp_pagado", "amount"),
    ("Saldo Insoluto", "imp_saldo_insoluto", "amount"),
    ("Archivo XML", "archivo_xml", "str"),
]


def export_to_xlsx(
    rows: list[AnyRow],
    output_path: Path | str,
    target_rfc: str | None = None,
) -> Path:
    """Write rows to a multi-sheet .xlsx file. Returns the output Path.

    When *target_rfc* is given the workbook uses an Emitidos/Recibidos split
    relative to that RFC. When it is None (Todos) the classic
    Ingresos/Egresos/Traslados structure is used instead.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if target_rfc:
        _build_filtered_workbook(wb, rows, target_rfc)
    else:
        _build_unfiltered_workbook(wb, rows)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def _build_filtered_workbook(wb, rows: list[AnyRow], target_rfc: str) -> None:
    """Emitidos/Recibidos structure relative to target_rfc."""
    emitidos = [r for r in rows if isinstance(r, CfdiRow) and r.rfc_emisor == target_rfc]
    recibidos = [r for r in rows if isinstance(r, CfdiRow) and r.rfc_receptor == target_rfc]
    nominas = [
        r
        for r in rows
        if isinstance(r, NominaRow) and (r.rfc_emisor == target_rfc or r.rfc_empleado == target_rfc)
    ]
    pagos = [r for r in rows if isinstance(r, PagoDocRow) and r.rfc_emisor == target_rfc]

    if emitidos:
        _write_data_sheet(wb.create_sheet("Emitidos"), emitidos, _FACTURA_COLS)
    if recibidos:
        _write_data_sheet(wb.create_sheet("Recibidos"), recibidos, _FACTURA_COLS)
    if nominas:
        _write_data_sheet(wb.create_sheet("Nóminas"), nominas, _NOMINA_COLS)
    if pagos:
        _write_data_sheet(wb.create_sheet("Complementos Pago"), pagos, _PAGO_COLS)

    _write_resumen_filtered_sheet(
        wb.create_sheet("Resumen"), emitidos, recibidos, nominas, pagos, target_rfc
    )


def _build_unfiltered_workbook(wb, rows: list[AnyRow]) -> None:
    """Classic Ingresos/Egresos/Traslados structure (no RFC filter)."""
    ingresos = [r for r in rows if isinstance(r, CfdiRow) and r.tipo_code == "I"]
    egresos = [r for r in rows if isinstance(r, CfdiRow) and r.tipo_code == "E"]
    traslados = [r for r in rows if isinstance(r, CfdiRow) and r.tipo_code == "T"]
    otros_cfdi = [r for r in rows if isinstance(r, CfdiRow) and r.tipo_code not in ("I", "E", "T")]
    nominas = [r for r in rows if isinstance(r, NominaRow)]
    pagos = [r for r in rows if isinstance(r, PagoDocRow)]

    if ingresos:
        _write_data_sheet(wb.create_sheet("Ingresos"), ingresos, _FACTURA_COLS)
    if egresos:
        _write_data_sheet(wb.create_sheet("Egresos"), egresos, _FACTURA_COLS)
    if traslados:
        _write_data_sheet(wb.create_sheet("Traslados"), traslados, _FACTURA_COLS)
    if nominas:
        _write_data_sheet(wb.create_sheet("Nóminas"), nominas, _NOMINA_COLS)
    if pagos:
        _write_data_sheet(wb.create_sheet("Complementos Pago"), pagos, _PAGO_COLS)
    if otros_cfdi:
        _write_data_sheet(wb.create_sheet("Otros"), otros_cfdi, _FACTURA_COLS)

    _write_resumen_sheet(wb.create_sheet("Resumen"), rows)


def _get_val(row: AnyRow, field: str):
    val = getattr(row, field, None)
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, datetime):
        return val.date()
    return val


def _write_header(ws: Worksheet, cols: list[_ColSpec]) -> None:
    for col_idx, (label, _, _) in enumerate(cols, 1):
        cell = ws.cell(1, col_idx, label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_cell_format(cell, fmt: str) -> None:
    if fmt == "amount":
        cell.number_format = _AMOUNT_FMT
    elif fmt == "date" and cell.value is not None:
        cell.number_format = _DATE_FMT
    elif fmt == "rate":
        cell.number_format = _RATE_FMT


def _write_data_sheet(ws: Worksheet, rows: list, cols: list[_ColSpec]) -> None:
    _write_header(ws, cols)

    amount_col_indices = {i + 1 for i, (_, _, fmt) in enumerate(cols) if fmt == "amount"}
    amount_sums: dict[int, float] = defaultdict(float)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, (_, field, fmt) in enumerate(cols, 1):
            val = _get_val(row, field)
            cell = ws.cell(row_idx, col_idx, val)
            _apply_cell_format(cell, fmt)
            if fmt == "amount" and isinstance(val, int | float):
                amount_sums[col_idx] += val

    # Totals row
    total_row = len(rows) + 2
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(total_row, col_idx)
        if col_idx == 1:
            cell.value = "TOTAL"
        elif col_idx in amount_col_indices:
            cell.value = amount_sums[col_idx]
            cell.number_format = _AMOUNT_FMT
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL

    _autofit_columns(ws)


def _write_resumen_sheet(ws: Worksheet, rows: list[AnyRow]) -> None:
    """One row per (RFC Emisor, Mes/Año, Tipo) aggregate."""
    agg: dict[tuple, dict] = defaultdict(
        lambda: {"num": 0, "subtotal": 0.0, "iva": 0.0, "total_mxn": 0.0}
    )

    for row in rows:
        if isinstance(row, CfdiRow):
            mes = row.fecha.strftime("%Y-%m") if row.fecha else "Sin fecha"
            key = (row.rfc_emisor, row.nombre_emisor, mes, row.tipo)
            agg[key]["num"] += 1
            agg[key]["subtotal"] += float(row.subtotal)
            agg[key]["iva"] += float(row.iva_trasladado)
            agg[key]["total_mxn"] += float(row.total_mxn)
        elif isinstance(row, NominaRow):
            mes = row.fecha.strftime("%Y-%m") if row.fecha else "Sin fecha"
            key = (row.rfc_emisor, row.nombre_emisor, mes, "Nómina")
            agg[key]["num"] += 1
            agg[key]["subtotal"] += float(row.total)
            agg[key]["total_mxn"] += float(row.total)
        elif isinstance(row, PagoDocRow):
            mes = row.fecha_pago.strftime("%Y-%m") if row.fecha_pago else "Sin fecha"
            key = (row.rfc_emisor, row.nombre_emisor, mes, "Pago")
            agg[key]["num"] += 1
            agg[key]["total_mxn"] += float(row.imp_pagado)

    headers = [
        "RFC Emisor",
        "Nombre Emisor",
        "Mes/Año",
        "Tipo",
        "Num. CFDIs",
        "Subtotal",
        "IVA",
        "Total MXN",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for row_idx, (key, vals) in enumerate(sorted(agg.items()), 2):
        rfc, nombre, mes, tipo = key
        ws.cell(row_idx, 1, rfc)
        ws.cell(row_idx, 2, nombre)
        ws.cell(row_idx, 3, mes)
        ws.cell(row_idx, 4, tipo)
        ws.cell(row_idx, 5, vals["num"])
        for col_idx, val_key in zip((6, 7, 8), ("subtotal", "iva", "total_mxn"), strict=False):
            cell = ws.cell(row_idx, col_idx, vals[val_key])
            cell.number_format = _AMOUNT_FMT

    _autofit_columns(ws)


def _write_resumen_filtered_sheet(
    ws: Worksheet,
    emitidos: list[CfdiRow],
    recibidos: list[CfdiRow],
    nominas: list[NominaRow],
    pagos: list[PagoDocRow],
    target_rfc: str = "",
) -> None:
    """Resumen aggregated by (Mes/Año, Dirección, Tipo) for a filtered workbook."""
    agg: dict[tuple, dict] = defaultdict(
        lambda: {"num": 0, "subtotal": 0.0, "iva": 0.0, "total_mxn": 0.0}
    )

    for row in emitidos:
        mes = row.fecha.strftime("%Y-%m") if row.fecha else "Sin fecha"
        key = (mes, "Emitido", row.tipo)
        agg[key]["num"] += 1
        agg[key]["subtotal"] += float(row.subtotal)
        agg[key]["iva"] += float(row.iva_trasladado)
        agg[key]["total_mxn"] += float(row.total_mxn)

    for row in recibidos:
        mes = row.fecha.strftime("%Y-%m") if row.fecha else "Sin fecha"
        key = (mes, "Recibido", row.tipo)
        agg[key]["num"] += 1
        agg[key]["subtotal"] += float(row.subtotal)
        agg[key]["iva"] += float(row.iva_trasladado)
        agg[key]["total_mxn"] += float(row.total_mxn)

    for row in nominas:
        mes = row.fecha.strftime("%Y-%m") if row.fecha else "Sin fecha"
        direccion = "Emitido" if row.rfc_emisor == target_rfc else "Recibido"
        key = (mes, direccion, "Nómina")
        agg[key]["num"] += 1
        agg[key]["subtotal"] += float(row.total)
        agg[key]["total_mxn"] += float(row.total)

    for row in pagos:
        mes = row.fecha_pago.strftime("%Y-%m") if row.fecha_pago else "Sin fecha"
        key = (mes, "Emitido", "Pago")
        agg[key]["num"] += 1
        agg[key]["total_mxn"] += float(row.imp_pagado)

    headers = ["Mes/Año", "Dirección", "Tipo", "Num. CFDIs", "Subtotal", "IVA", "Total MXN"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for row_idx, (key, vals) in enumerate(sorted(agg.items()), 2):
        mes, direccion, tipo = key
        ws.cell(row_idx, 1, mes)
        ws.cell(row_idx, 2, direccion)
        ws.cell(row_idx, 3, tipo)
        ws.cell(row_idx, 4, vals["num"])
        for col_idx, val_key in zip((5, 6, 7), ("subtotal", "iva", "total_mxn"), strict=False):
            cell = ws.cell(row_idx, col_idx, vals[val_key])
            cell.number_format = _AMOUNT_FMT

    _autofit_columns(ws)


def _autofit_columns(ws: Worksheet) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
