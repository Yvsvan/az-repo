"""Tests TDD para el exportador Excel.

Estructura esperada del .xlsx:
  - Una hoja por RFC con los movimientos (nombre = RFC o "SIN_RFC_N").
  - Encabezados en negrita en la fila 1.
  - Fila de TOTAL al pie con suma de abonos y cargos.
  - Hoja "Resumen" con una fila por Statement.
  - Hoja "Advertencias" sólo si hay warnings.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from bank_parser.core.schema import Statement
from bank_parser.exporters.excel import COLUMNAS, export_to_xlsx

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _load(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path)


def _sheet_names(path: Path) -> list[str]:
    return _load(path).sheetnames


def _header_row(ws) -> list[str]:
    return [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]


def _data_rows(ws) -> list[list]:
    """Filas 2..max (sin encabezado)."""
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
    return rows


# ---------------------------------------------------------------------------
# retorno de path
# ---------------------------------------------------------------------------


def test_returns_path_object(tmp_path: Path, sample_statement: Statement) -> None:
    result = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    assert isinstance(result, Path)
    assert result.exists()


def test_creates_file_at_given_path(tmp_path: Path, sample_statement: Statement) -> None:
    dest = tmp_path / "sub" / "result.xlsx"
    export_to_xlsx([sample_statement], dest)
    assert dest.exists()


# ---------------------------------------------------------------------------
# hojas por RFC
# ---------------------------------------------------------------------------


def test_single_rfc_creates_one_movement_sheet(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    rfc = sample_statement.summary.rfc
    assert rfc in _sheet_names(path)


def test_two_rfcs_create_two_movement_sheets(
    tmp_path: Path, two_statements_diff_rfc: list[Statement]
) -> None:
    path = export_to_xlsx(two_statements_diff_rfc, tmp_path / "out.xlsx")
    names = _sheet_names(path)
    rfcs = {s.summary.rfc for s in two_statements_diff_rfc}
    for rfc in rfcs:
        assert rfc in names


def test_same_rfc_two_statements_one_sheet(
    tmp_path: Path, two_statements_same_rfc: list[Statement]
) -> None:
    path = export_to_xlsx(two_statements_same_rfc, tmp_path / "out.xlsx")
    rfc = two_statements_same_rfc[0].summary.rfc
    names = _sheet_names(path)
    # Solo una hoja con ese RFC (no duplica)
    assert names.count(rfc) == 1


def test_no_rfc_sheet_named_sin_rfc(tmp_path: Path, statement_no_rfc: Statement) -> None:
    path = export_to_xlsx([statement_no_rfc], tmp_path / "out.xlsx")
    names = _sheet_names(path)
    assert any("SIN_RFC" in n for n in names)


# ---------------------------------------------------------------------------
# columnas estándar
# ---------------------------------------------------------------------------


def test_standard_columns_present(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    rfc = sample_statement.summary.rfc
    ws = wb[rfc]
    headers = _header_row(ws)
    for col in COLUMNAS:
        assert col in headers, f"Columna '{col}' faltante"


def test_columns_in_exact_order(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb[sample_statement.summary.rfc]
    assert _header_row(ws) == list(COLUMNAS)


# ---------------------------------------------------------------------------
# datos
# ---------------------------------------------------------------------------


def test_movement_count_equals_data_rows_plus_total(
    tmp_path: Path, sample_statement: Statement
) -> None:
    """Las filas de datos = len(movements) + 1 fila TOTAL."""
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb[sample_statement.summary.rfc]
    # max_row = 1 encabezado + N movimientos + 1 TOTAL
    expected = 1 + len(sample_statement.movements) + 1
    assert ws.max_row == expected


def test_same_rfc_consolidates_movements(
    tmp_path: Path, two_statements_same_rfc: list[Statement]
) -> None:
    total_movs = sum(len(s.movements) for s in two_statements_same_rfc)
    path = export_to_xlsx(two_statements_same_rfc, tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb[two_statements_same_rfc[0].summary.rfc]
    # max_row = 1 + total_movs + 1 (TOTAL)
    assert ws.max_row == 1 + total_movs + 1


def test_amounts_stored_as_numbers(tmp_path: Path, sample_statement: Statement) -> None:
    """abono, cargo y saldo deben ser numéricos (float/int), no strings."""
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb[sample_statement.summary.rfc]
    abono_col = list(COLUMNAS).index("abono") + 1
    cargo_col = list(COLUMNAS).index("cargo") + 1
    saldo_col = list(COLUMNAS).index("saldo") + 1
    for row in range(2, ws.max_row):  # excluye fila TOTAL que puede ser fórmula/número
        assert isinstance(ws.cell(row, abono_col).value, int | float), (
            f"abono fila {row} no es número"
        )
        assert isinstance(ws.cell(row, cargo_col).value, int | float), (
            f"cargo fila {row} no es número"
        )
        assert isinstance(ws.cell(row, saldo_col).value, int | float), (
            f"saldo fila {row} no es número"
        )


def test_fecha_stored_as_date_or_string(tmp_path: Path, sample_statement: Statement) -> None:
    """fecha debe ser date o string ISO, no None."""
    from datetime import date as date_type

    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb[sample_statement.summary.rfc]
    fecha_col = list(COLUMNAS).index("fecha") + 1
    for row in range(2, ws.max_row):  # excluye fila TOTAL
        val = ws.cell(row, fecha_col).value
        assert val is not None
        assert isinstance(val, date_type | str)


# ---------------------------------------------------------------------------
# fila de totales
# ---------------------------------------------------------------------------


def test_total_row_label_in_descripcion(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb[sample_statement.summary.rfc]
    desc_col = list(COLUMNAS).index("descripcion") + 1
    last_row_desc = ws.cell(ws.max_row, desc_col).value
    assert str(last_row_desc).upper() == "TOTAL"


def test_total_row_abono_sum(tmp_path: Path, sample_statement: Statement) -> None:
    expected = float(sum(m.abono for m in sample_statement.movements))
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb[sample_statement.summary.rfc]
    abono_col = list(COLUMNAS).index("abono") + 1
    total_val = ws.cell(ws.max_row, abono_col).value
    assert abs(float(total_val) - expected) < 0.01


def test_total_row_cargo_sum(tmp_path: Path, sample_statement: Statement) -> None:
    expected = float(sum(m.cargo for m in sample_statement.movements))
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb[sample_statement.summary.rfc]
    cargo_col = list(COLUMNAS).index("cargo") + 1
    total_val = ws.cell(ws.max_row, cargo_col).value
    assert abs(float(total_val) - expected) < 0.01


# ---------------------------------------------------------------------------
# hoja Resumen
# ---------------------------------------------------------------------------


def test_resumen_sheet_exists(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    assert "Resumen" in _sheet_names(path)


def test_resumen_sheet_has_one_row_per_statement(
    tmp_path: Path, two_statements_diff_rfc: list[Statement]
) -> None:
    path = export_to_xlsx(two_statements_diff_rfc, tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb["Resumen"]
    # 1 encabezado + N statements
    assert ws.max_row == 1 + len(two_statements_diff_rfc)


def test_resumen_contains_rfc_column(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb["Resumen"]
    headers = _header_row(ws)
    assert "rfc" in headers


def test_resumen_contains_saldo_columns(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb["Resumen"]
    headers = _header_row(ws)
    assert "saldo_inicial" in headers
    assert "saldo_final" in headers


# ---------------------------------------------------------------------------
# hoja Advertencias
# ---------------------------------------------------------------------------


def test_no_warnings_sheet_when_no_warnings(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    assert "Advertencias" not in _sheet_names(path)


def test_warnings_sheet_exists_when_warnings(
    tmp_path: Path, statement_with_warnings: Statement
) -> None:
    path = export_to_xlsx([statement_with_warnings], tmp_path / "out.xlsx")
    assert "Advertencias" in _sheet_names(path)


def test_warnings_sheet_row_count(tmp_path: Path, statement_with_warnings: Statement) -> None:
    path = export_to_xlsx([statement_with_warnings], tmp_path / "out.xlsx")
    wb = _load(path)
    ws = wb["Advertencias"]
    total_warnings = len(statement_with_warnings.warnings)
    # 1 encabezado + N warnings
    assert ws.max_row == 1 + total_warnings


# ---------------------------------------------------------------------------
# sheet order: movimientos primero, Resumen, Advertencias al final
# ---------------------------------------------------------------------------


def test_resumen_is_last_or_near_last(tmp_path: Path, sample_statement: Statement) -> None:
    path = export_to_xlsx([sample_statement], tmp_path / "out.xlsx")
    names = _sheet_names(path)
    # Resumen debe aparecer después de todas las hojas de movimientos
    rfc_idx = names.index(sample_statement.summary.rfc)
    resumen_idx = names.index("Resumen")
    assert resumen_idx > rfc_idx
